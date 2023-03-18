import os
from time import sleep

from selene import browser
from selenium import webdriver
from selenium.webdriver.chrome.service import Service as ChromeService
from webdriver_manager.chrome import ChromeDriverManager

from PyPDF2 import PdfReader
from openpyxl import load_workbook
from zipfile import ZipFile

url = 'https://filesamples.com/formats/'


def row_xlsx_file():
    row_in_xlsx = load_workbook('sample1.xlsx').active.max_row
    return row_in_xlsx


def row_csv_file():
    with open('sample4.csv') as file:
        row_in_csv = 0
        for row in file:
            row_in_csv += 1

    return row_in_csv


def page_pdf_file():
    with open('sample2.pdf') as file:
        pdf_reader = PdfReader('sample2.pdf')
        page_in_pdf = len(pdf_reader.pages)

    return page_in_pdf


def size_file():
    size_csv = os.path.getsize('sample4.csv')
    size_pdf = os.path.getsize('sample2.pdf')
    size_xlsx = os.path.getsize('sample1.xlsx')
    return size_csv, size_pdf, size_xlsx


def test_ar():
    browser.config.window_height = 1800
    browser.config.window_width = 1800
    browser.config.hold_browser_open = True
    options = webdriver.ChromeOptions()

    prefs = {
        "download.default_directory": "D:\\lesson_7\\test",
        "download.prompt_for_download": False
    }

    options.add_experimental_option('prefs', prefs)
    driver = webdriver.Chrome(service=ChromeService(ChromeDriverManager().install()), options=options)

    browser.config.driver = driver

    # путь до test_archive.py
    py_file = os.path.abspath(__file__)
    # путь до dir test
    test_dir = os.path.dirname(py_file)
    # путь до dir lesson_7, с добавлением новой dir resources
    resources_dir = os.path.abspath(os.path.join(test_dir, '..', 'resources'))
    # lesson_7_dir = os.path.abspath(os.path.join(test_dir, '..'))

    if not os.path.exists(test_dir + '\\sample4.csv'):
        browser.open(url + 'csv')
        browser.element(css_or_xpath_or_by="(//a[contains(text(), 'Download')])[1]").click()
        sleep(5)

    if not os.path.exists(test_dir + '\\sample2.pdf'):
        browser.open(url + 'pdf')
        browser.element(css_or_xpath_or_by="(//a[contains(text(), 'Download')])[2]").click()
        sleep(5)

    if not os.path.exists(test_dir + '\\sample1.xlsx'):
        browser.open(url + 'xlsx')
        browser.element(css_or_xpath_or_by="(//a[contains(text(), 'Download')])[3]").click()
        sleep(5)

    if driver.session_id:
        browser.close()

    if not os.path.exists(resources_dir):
        # создание dir resources
        os.mkdir(resources_dir)

    if os.path.exists(test_dir + '\\test_ar.zip'):
        os.remove(test_dir + '\\test_ar.zip')

    with ZipFile("test_ar.zip", 'a') as myzip:
        myzip.write('sample1.xlsx')
        myzip.write('sample2.pdf')
        myzip.write('sample4.csv')

    os.replace(test_dir + '\\test_ar.zip', resources_dir + '\\test_ar.zip')

    with ZipFile(resources_dir + "//test_ar.zip") as myzip:
        csv_zip_1 = myzip.getinfo('sample4.csv').file_size
        pdf_zip_1 = myzip.getinfo('sample2.pdf').file_size
        xlsx_zip_1 = myzip.getinfo('sample1.xlsx').file_size

        with myzip.open('sample4.csv') as file:
            csv_zip_2 = 0
            for row in file:
                csv_zip_2 += 1

        with myzip.open('sample2.pdf') as file:
            pdf_reader = PdfReader(file)
            pdf_zip_2 = len(pdf_reader.pages)

        with myzip.open('sample1.xlsx') as file:
            xlsx_zip_2 = load_workbook(file).active.max_row

    size_csv, size_pdf, size_xlsx = size_file()

    assert csv_zip_1 == size_csv
    assert pdf_zip_1 == size_pdf
    assert xlsx_zip_1 == size_xlsx

    assert csv_zip_2 == row_csv_file()
    assert pdf_zip_2 == page_pdf_file()
    assert xlsx_zip_2 == row_xlsx_file()
